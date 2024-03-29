from pathlib import Path
import openpyxl
import xlwt
import os

# 设置工作目录到脚本所在目录
os.chdir(os.path.abspath(os.path.dirname(__file__)))

# 指定输出目录
pdir = Path('output')

# 获取所有.xlsx文件
filelist = [filename for filename in pdir.iterdir() if filename.suffix == '.xlsx']

for infile in filelist:
    # 使用openpyxl加载.xlsx文件
    workbook = openpyxl.load_workbook(infile)
    sheet = workbook.active

    # 创建一个新的xlwt工作簿和工作表
    new_workbook = xlwt.Workbook()
    new_sheet = new_workbook.add_sheet(sheet.title)

    # 复制数据
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        for col_idx, value in enumerate(row):
            new_sheet.write(row_idx, col_idx, value)

    # 构建输出文件名
    outfile = infile.with_suffix('.xls')

    # 保存为.xls格式
    new_workbook.save(outfile)

    print(f"Converted and saved: {outfile}")
